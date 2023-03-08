#!/usr/bin/env python
# 
# For NSF API docs, see:
# https://www.research.gov/common/webapi/awardapisearch-v1.htm
#
import argparse
import datetime
import logging
import sys
import json
import requests
from requests.adapters import HTTPAdapter
from openpyxl import load_workbook
import xlsxwriter


logging.basicConfig(level=logging.INFO)

BASE_URL='http://api.nsf.gov/services/v1/'
SEARCH_URL=BASE_URL + 'awards.json'
RETRIEVE_URL=BASE_URL + 'awards/'
AWARD_INFO=['id',
            'agency',
            'awardeeName',
            'startDate',
            'expDate',
            'estimatedTotalAmt',
            'piFirstName',
            'piLastName',
            'pdPIName',
            'coPDPI',
            'title'
           ]


def search_by_date_range(start, end, institution):
    """
    Search by a range of dates and award institution name. Returns a list of award
    IDs matching the institution and that started within the range of dates.

    Date format: 'mm/dd/yyyy'
    Institution format: '"Name+of+Institution"'
    """

    offset_value = 1
    award_id_list = []

    while True:
        logging.info(f'searching with offset_value = {offset_value}')
        query_parameters = ''.join([ '?', 'startDateStart=', start, 
                                     '&', 'startDateEnd=', end,
                                     '&', 'awardeeName="', institution, '"',
                                     '&', 'offset=', str(offset_value) ])
        
        http = requests.Session()
        adapter = HTTPAdapter(max_retries=5)
        adapter.max_retries.respec_retry_after_header = False
        http.mount('http://', adapter)
        query_url = ''.join([SEARCH_URL, query_parameters])
        logging.info(f'getting {query_url}')
        try:
            response = http.get(query_url, timeout=20)
        except requests.exceptions.ReadTimeout:
            print('timeout during search...try again later')
            sys.exit()
        except Exception as x:
            print(f'request failed because {x}')
            sys.exit()

        logging.info(f'response was {response.ok}')

        for i in range(len(response.json()['response']['award'])):
            award_id_list.append(response.json()['response']['award'][i]['id'])

        logging.debug(award_id_list)
        if len(response.json()['response']['award']) == 25:
            offset_value += 25
            continue
        else:
            logging.info(f'{len(award_id_list)} awards found')
            return award_id_list

    
def retrieve_award_info(award_id_list):
    """
    Retrieve specific award information given an award ID. The information returned
    is in the AWARD_INFO list.
    """

    award_dict = {}
    print_fields = ','.join(AWARD_INFO)

    for item in award_id_list:
        http = requests.Session()
        adapter = HTTPAdapter(max_retries=5)
        adapter.max_retries.respec_retry_after_header = False
        http.mount('http://', adapter)
        query_parameters = ''.join(['?', 'printFields=', print_fields])
        logging.info(f'getting {RETRIEVE_URL}{item}.json{query_parameters}')
        try:
            response = http.get(url=RETRIEVE_URL + item + '.json' + query_parameters, timeout=20)
        except requests.exceptions.ReadTimeout:
            print('timeout during award lookup...try again later')
            sys.exit()
        except Exception as x:
            print(f'request failed because {x}')
            sys.exit()

        logging.info(f'response was {response.ok}')

        award_dict[item] = response.json()['response']['award'][0]

        for field in AWARD_INFO:
            if field not in award_dict[item]:
                award_dict[item][field] = 'NO DATA AVAILABLE' 
    return award_dict


def write_output_sheet(award_dict, userlist, output):
    """
    Given a dictionary of award information and a list of TACC usernames, write
    an output workbook with two worksheets: (1) Awards that match a TACC username
    and (2) awards that don't match a TACC username.
    """

    userlist_wb = load_workbook(filename=userlist, read_only=True)
    worksheet = userlist_wb['utrc_institution_accounts']
    row_count = worksheet.max_row
    rows = worksheet.rows

    name_dict = {}

    if row_count > 1:
        next(rows) # skip header row
        for row in rows:
            institution = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            name = ' '.join([first_name, last_name]).lower().replace(' ','')
            name_dict[name] = [institution, first_name, last_name]

    logging.info(f'number of items in name_dict = {len(name_dict.keys())}')

    workbook = xlsxwriter.Workbook(output)
    bold = workbook.add_format({'bold': 1})
    found_worksheet = workbook.add_worksheet('utrc_nsf_funding')
    found_worksheet.write_row(0, 0, ['utrc_institution', 'utrc_first_name', 'utrc_last_name']+AWARD_INFO, bold)
    not_found_worksheet = workbook.add_worksheet('not_utrc_nsf_funding')
    not_found_worksheet.write_row(0, 0, AWARD_INFO, bold)

    f_row = 1
    nf_row = 1
    for item in award_dict.keys():
        name_str = award_dict[item]['pdPIName'].lower().replace(' ','')
        if name_str in name_dict.keys():
            logging.info(f'{name_str} matches {name_dict[name_str]}')
            found_worksheet.write_row(f_row, 0, [name_dict[name_str][0],
                                                 name_dict[name_str][1],
                                                 name_dict[name_str][2],
                                                 award_dict[item]['id'],
                                                 award_dict[item]['agency'],
                                                 award_dict[item]['awardeeName'],
                                                 award_dict[item]['startDate'],
                                                 award_dict[item]['expDate'],
                                                 award_dict[item]['estimatedTotalAmt'],
                                                 award_dict[item]['piFirstName'],
                                                 award_dict[item]['piLastName'],
                                                 award_dict[item]['pdPIName'],
                                                 json.dumps(award_dict[item]['coPDPI']),
                                                 award_dict[item]['title']
                                                ])
            f_row += 1
        else:
            logging.info(f'{name_str} has no match')
            not_found_worksheet.write_row(nf_row, 0, [award_dict[item]['id'],
                                                      award_dict[item]['agency'],
                                                      award_dict[item]['awardeeName'],
                                                      award_dict[item]['startDate'],
                                                      award_dict[item]['expDate'],
                                                      award_dict[item]['estimatedTotalAmt'],
                                                      award_dict[item]['piFirstName'],
                                                      award_dict[item]['piLastName'],
                                                      award_dict[item]['pdPIName'],
                                                      json.dumps(award_dict[item]['coPDPI']),
                                                      award_dict[item]['title']
                                                     ])
            nf_row += 1

    workbook.close()
    return


def main():

    parser = argparse.ArgumentParser(description='Scrape NSF funded awards')
    parser.add_argument('-s', '--start', dest='start_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-e', '--end', dest='end_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-i', '--institution', dest='inst', help='institution search term, format = University+of+Texas', required=True)
    parser.add_argument('-u', '--userlist', dest='userlist', help='input file with list of names and affiliations', required=True)
    parser.add_argument('-o', '--output', dest='output', help='output file', required=True)
    args = parser.parse_args()

    start = datetime.datetime.strptime(args.start_date, '%Y%m%d').strftime('%m/%d/%Y')
    end = datetime.datetime.strptime(args.end_date, '%Y%m%d').strftime('%m/%d/%Y')

    award_id_list = search_by_date_range(start, end, args.inst)
    award_dict = retrieve_award_info(award_id_list)
    write_output_sheet(award_dict, '/data/' + args.userlist, '/data/' + f'NSF_{args.output}')

    return


if __name__ == '__main__':
    main()

