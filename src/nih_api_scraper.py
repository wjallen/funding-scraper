#!/usr/bin/env python
# 
# For NIH API docs, see:
# https://api.reporter.nih.gov/
#

import argparse
from datetime import datetime, timedelta
import logging
import json
import requests
from openpyxl import load_workbook
import xlsxwriter
from fuzzywuzzy import fuzz
import math

logging.basicConfig(level=logging.DEBUG,
format='%(asctime)s %(levelname)s %(message)s',
      filename='/data/nih.log',
      filemode='w')

"""
Define the url for NIH API and the info being
saved to our spreadsheets. AWARD_INFO is used to
format our sheets that hold the desired data. 
ALL_UNIVERSITIES maps the API institution names 
to our own. 
"""

URL = 'https://api.reporter.nih.gov/v2/projects/search'

AWARD_INFO=['id',
            'agency',
            'awardeeName',
            'startDate',
            'expDate',
            'estimatedTotalAmt',
            'piFirstName',
            'piLastName',
            'pdPIName',
            'title',
            'coPDPI',
            'taccPDPI',
           ]

ALL_UNIVERSITIES={
    'UNIVERSITY OF TEXAS RIO GRANDE VALLEY': 'University of Texas Rio Grande Valley', 
    'UNIVERSITY OF TEXAS DALLAS': 'University of Texas at Dallas (UTD) (UT Dallas)', 
    'UNIVERSITY OF TEXAS HLTH CTR AT TYLER': 'University of Texas Health Science Center at Tyler', 
    'UNIVERSITY OF TX MD ANDERSON CAN CTR': 'University of Texas MD Anderson Cancer Center', 
    'UNIVERSITY OF TEXAS AT AUSTIN': 'University of Texas at Austin (UT) (UT Austin)', 
    'UNIVERSITY OF TEXAS HLTH SCI CTR HOUSTON': 'University of Texas Health Science Center at Houston', 
    'UT SOUTHWESTERN MEDICAL CENTER': 'University of Texas Southwestern Medical Center (UTSW) (UT Southwestern)', 
    'UNIVERSITY OF TEXAS TYLER': 'University of Texas Tyler', 
    'UNIVERSITY OF TEXAS HLTH SCIENCE CENTER': 'University of Texas Health Science Center at San Antonio', 
    'UNIVERSITY OF TEXAS ARLINGTON': 'University of Texas at Arlington (UTA) (UT Arlington)', 
    'UNIVERSITY OF TEXAS EL PASO': 'University of Texas at El Paso (UTEP)', 
    'UNIVERSITY OF TEXAS MED BR GALVESTON': 'University of Texas Medical Branch at Galveston', 
    'UNIVERSITY OF TEXAS OF THE PERMIAN BASIN': 'University of Texas Permian Basin', 
    'UNIVERSITY OF TEXAS SAN ANTONIO': 'University of Texas at San Antonio'
}

def partition(l, n):
    for i in range(0, len(l), n):
        yield l[i:i + n]

def splitDateRange(origin,N,blocks):

    """
    Given an origin (start date), the number of chunks to make,
    and the amount of days stored in a list of blocks for each chunk, the
    following function divides the users date range to combat the 500 response
    limit. The max amount of days in a range per request is 75.
    """

    api_calls = []
    timeDict ={}
    timeList= []
    timeList.append(origin)
    for x in range(1,N):
        timeList.append(timeList[x-1] + timedelta(days=len(blocks[x-1])))

    timeDict[0] = [timeList[0], timeList[0]+ timedelta(days=len(blocks[0])-1)]
    for x in range(1,N):
        timeDict[x] = [timeList[x] -timedelta(days=1), (timeList[x]) + timedelta(days=len(blocks[x])-1)]

    timeDict[len(timeDict)-1] = [timeDict[len(timeDict)-1][0], (timeDict[len(timeDict)-1][1]+timedelta(days=1))]
    
    for x in timeDict:
        texas = {
            "criteria":
            {
                "project_start_date": { "from_date": str(timeDict[x][0].date()), "to_date": str(timeDict[x][1].date()) },
                "org_names": ["UNIVERSITY OF TEXAS","University of TX","UT SOUTHWESTERN MEDICAL CENTER"]
            },
                "limit": 500,
                "offset":0,
                "sort_field":"project_start_date",
                "sort_order":"desc"
            }
        api_calls.append(texas)

    return api_calls

def findAllProjects(start,end,calls):

    """
    Given a start date, end date, and a list of json payloads the function makes 
    a POST call to NIH API for each payload. Each payload is given a date range and a 
    list of strings for the query. Data from the response is parsed and appended to our 
    list of formatted objects. North Texas results are removed.
    """

    all_results = []
    results = []

    for x in calls:
        response = requests.post(URL, json = x).json()
        temp = response["results"]
        assert(len(temp) < 500), "The date range provided too many results, please provide a block smaller than 75 days."
        results += temp

    for y in results:
        if(results.count(y) > 1):
            assert("Duplicates in the response")
    
    logging.info(f"START: {start} END: {end}")
    print(f'Before removing North Texas: {len(results)}')

    for x in results:

        # Remove all of North Texas Results

        if ('NORTH' in x['organization']['org_name']):
            logging.info(f"REMOVING: {x['organization']['org_name']}")
            continue

        startDate = x["project_start_date"]
        if(startDate):
            startDate = startDate[5:7] + "/" + startDate[8:10] + "/" + startDate[0:4]
        endDate = x["project_end_date"]
        if(endDate):
            endDate = endDate[5:7] + "/" + endDate[8:10] + "/" + endDate[0:4]
        else:
            endDate = "N/A"

        coPDPI = []
        for y in x["principal_investigators"]:
            if(y["is_contact_pi"] == True):
                piFirstName = y["first_name"]
                piLastName = y["last_name"]
            else:
                coPDPI.append({
                    "first_name": y["first_name"],
                    "middle_name": y["last_name"],
                    "last_name" : y["last_name"]
                })

        # Create objects with the information we want.

        myObj = {
            "id" : x["appl_id"],
            "agency": x["agency_ic_fundings"][0]["abbreviation"],
            "awardeeName": x["organization"]["org_name"],
            "piFirstName": piFirstName,
            "piLastName": piLastName,
            "coPDPI": coPDPI or "NO DATA AVAILABLE",
            "pdPIName": x["contact_pi_name"],
            "startDate": startDate,
            "expDate": endDate,
            "estimatedTotalAmt": x["award_amount"],
            "title": x["project_title"],
            "city" : x["organization"]["org_city"]
        }

        all_results.append(myObj)

    print(f'After removing North Texas: {len(all_results)}')
    return all_results

def findTACCUsers(userlist,output,awards):

    """
    Given a list of award information and a list of TACC usernames, write
    an output workbook with two worksheets: (1) Awards that match a TACC username
    and (2) awards that don't match a TACC username.
    """

    userlist_wb = load_workbook(filename=userlist, read_only=True)
    worksheet = userlist_wb['utrc_institution_accounts']
    row_count = worksheet.max_row
    rows = worksheet.rows

    name_dict = {}

    if row_count > 1:
        next(rows) # skip header row
        for row in rows:
            institution = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            name = ' '.join([first_name, last_name]).lower().replace(' ','')
            name_dict[name] = [institution, first_name, last_name]

    logging.info(f'number of items in name_dict = {len(name_dict.keys())}')

    workbook = xlsxwriter.Workbook(output)
    bold = workbook.add_format({'bold': 1})
    found_worksheet = workbook.add_worksheet('utrc_nih_funding')
    found_worksheet.write_row(0, 0, ['utrc_institution', 'utrc_first_name', 'utrc_last_name']+AWARD_INFO, bold)
    not_found_worksheet = workbook.add_worksheet('not_utrc_nih_funding')
    not_found_worksheet.write_row(0, 0, AWARD_INFO, bold)

    f_format = workbook.add_format({'bg_color':'#90EE90'})
    nf_format = workbook.add_format({'bg_color':'#FCC981'})

    f_row = 1
    nf_row = 1
    fuzzy_names = 0
    saved_names = 0

    for item in awards:    
        collab_format = workbook.add_format({'font_color':'red'})

        name_str = item['piFirstName'].lower() + item['piLastName'].lower()
        name_str = name_str.replace(" ", "")
        first_name_str = item['piFirstName'].lower()
        last_name_str = item['piLastName'].lower()
        affiliation = item['awardeeName']

        collaborators = []
        formattedCollab = []
        collab_str = ""

        if(item['coPDPI']!= "NO DATA AVAILABLE"):
            collaborators = item['coPDPI']

        # If a collaborator is in the TACC system, save it for proper formatting.
        # Check for fuzzywuzzy name matching on collaborators.
            
        if(collaborators):
            for z in collaborators:
                collab_str = z['first_name'].lower() + z['last_name'].lower()
                if collab_str in name_dict.keys():
                    formattedCollab.append(z['first_name'] + " " + z['last_name'])
                else:
                    for x in name_dict:
                        if(z['last_name'].lower() != name_dict[x][2].lower()):
                            continue
                        y = fuzz.ratio(z['first_name'].lower(),name_dict[x][1].lower())
                        if(y >= 89 and y < 100 ):
                            fuzzy_names += 1
                            saved_names += 1
                            logging.warning(f"Collaborator {z['first_name']} {z['last_name']} was found based on fuzzywuzzy ratio")
                            formattedCollab.append(name_dict[x][1] + " " + name_dict[x][2])
                            collab_format = workbook.add_format({'bg_color': '#90EE90', 'font_color' : 'red'})


        # If the name matches one in our TACC system, add it to the found sheet. 
        # If the collaborators are in our TACC systems, highlight their names red.

        if name_str in name_dict.keys():
            logging.info(f'{name_str} matches {name_dict[name_str]}')
            found_worksheet.write_row(f_row, 0, [name_dict[name_str][0],
                                                name_dict[name_str][1],
                                                name_dict[name_str][2],
                                                item['id'],
                                                item['agency'],
                                                item['awardeeName'],
                                                item['startDate'],
                                                item['expDate'],
                                                item['estimatedTotalAmt'],
                                                item['piFirstName'],
                                                item['piLastName'],
                                                item['pdPIName'],
                                                item['title'],
                                                json.dumps(item['coPDPI'])
                                                ])
            if(formattedCollab):
                found_worksheet.write(f_row,14,json.dumps(formattedCollab),collab_format)
            else:
                found_worksheet.write(f_row,14,"None Found")
            f_row += 1

        # If the name does not match one in our TACC system, but a collaborator does, add it to
        # the found sheet. Collaborator will be highlighted in red.

        elif formattedCollab:
            found_worksheet.write_row(f_row, 0, [name_dict[formattedCollab[0].lower().replace(" ","")][0],
                                                item['piFirstName'],
                                                item['piLastName'],
                                                item['id'],
                                                item['agency'],
                                                item['awardeeName'],
                                                item['startDate'],
                                                item['expDate'],
                                                item['estimatedTotalAmt'],
                                                item['piFirstName'],
                                                item['piLastName'],
                                                item['pdPIName'],
                                                item['title'],
                                                ])
            found_worksheet.write(f_row,13,json.dumps(item['coPDPI']),collab_format)
            found_worksheet.write(f_row,14,json.dumps(formattedCollab),collab_format)
            f_row += 1

        # If the name does not match one in our TACC system, we will search through names that have an exact 
        # last name match. The first name will be compared using fuzzywuzzy word matching. If this returns 
        # a score of 89 or higher, we will pass the PI as a match.

        else:
            logging.info(f'{name_str} has no match')
            
            fuzzy = False
            following = True
            added = False

            for x in name_dict:
                if(last_name_str != name_dict[x][2].lower()):
                    continue
                y = fuzz.ratio(first_name_str,name_dict[x][1].lower())
                if(y >= 80):
                    fuzzy_names += 1
                    fuzzy = True
                    logging.warning(f"Ratio of {y} for {first_name_str} {last_name_str} and {name_dict[x][1].lower()} {name_dict[x][2].lower()}")
                    logging.warning(f"PI Affiliation: {affiliation} && TACC User Affiliation: {name_dict[x][0]}")
                    if(y >= 89 and y < 100 ):
                        saved_names += 1
                        logging.warning(f"Moving {first_name_str} {last_name_str} into sheet (i) based on fuzzywuzzy ratio")
                        if  not added:
                            found_worksheet.write_row(f_row, 0, [name_dict[x][0],
                                                        name_dict[x][1],
                                                        name_dict[x][2],
                                                        item['id'],
                                                        item['agency'],
                                                        item['awardeeName'],
                                                        item['startDate'],
                                                        item['expDate'],
                                                        item['estimatedTotalAmt'],
                                                        item['piFirstName'],
                                                        item['piLastName'],
                                                        item['pdPIName'],
                                                        item['title'],
                                                        json.dumps(item['coPDPI']),
                                                        "None Found"
                                                        ],f_format)
                            f_row += 1
                            following = False
                            added = True
                    
                        
            if following:
                if fuzzy:
                    format = nf_format
                else:
                    format = None

                not_found_worksheet.write_row(nf_row, 0,[ item['id'],
                                                            item['agency'],
                                                            item['awardeeName'],
                                                            item['startDate'],
                                                            item['expDate'],
                                                            item['estimatedTotalAmt'],
                                                            item['piFirstName'],
                                                            item['piLastName'],
                                                            item['pdPIName'],
                                                            item['title'],
                                                            json.dumps(item['coPDPI']),
                                                            "None Found"
                                                        ], format)                                                       
                nf_row += 1

    found = f_row - 1
    notFound = nf_row - 1  
    if(notFound != 0):
        logging.info("TACC Percentage: {:.2f}".format(float(found/notFound) * 100) + "%")
    logging.info(f'Total fuzzy names: {fuzzy_names}' )
    logging.info(f'Fuzzy names saved: {saved_names}' )

    workbook.close()
    return


def main():

    parser = argparse.ArgumentParser(description='Scrape NSF funded awards')
    parser.add_argument('-s', '--start', dest='start_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-e', '--end', dest='end_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-i', '--institution', dest='inst', help='institution search term, format = University+of+Texas', required=True)
    parser.add_argument('-u', '--userlist', dest='userlist', help='input file with list of names and affiliations', required=True)
    parser.add_argument('-o', '--output', dest='output', help='output file', required=True)
    args = parser.parse_args()

    start = str(args.start_date)[0:4] + "-" + str(args.start_date)[4:6] + "-" + str(args.start_date)[6:]
    end = str(args.end_date)[0:4] + "-" + str(args.end_date)[4:6] + "-" + str(args.end_date)[6:]


    # parse arguments to split date range and avoid a request limit
    # using a 75 day split

    origin = datetime.strptime(start,"%Y-%m-%d")
    finish = datetime.strptime(end,"%Y-%m-%d")
    days = (finish - origin).days
    groups = math.ceil(days/75)
    l = list(range(0,days))
    n = math.ceil(days/groups)
    chunks = list(partition(l,n))
    
    # split user inputted date range, get all NIH awards, match TACC Users
    api_calls = []
    all_awards = []

    api_calls = splitDateRange(origin, groups, chunks)
    all_awards = findAllProjects(start,end,api_calls)
    findTACCUsers('/data/' + args.userlist, '/data/' + args.output, all_awards)

if __name__ == '__main__':
    main()