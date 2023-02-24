import requests
from bs4 import BeautifulSoup
import argparse
import datetime
import re
import logging
from openpyxl import load_workbook
import xlsxwriter
from fuzzywuzzy import fuzz
import sys

logging.basicConfig(level=logging.DEBUG)

AWARD_INFO = ['Award Number',
              'Title',
              'Institution',
              'PI First Name',
              'PI Last Name',
              'Org Code',
              'Program Office',
              'PM',
              'Start Date',
              'End Date',
              'Most Recent Award Date',
              'Award Type',
              'Amount Awarded to Date',
              'Amount Awarded this FY',
              'Institution Type',
              'UEI',
              'Program Area',
              'Register Number',
              'DUNS'
              ]


def main():
    url = 'https://pamspublic.science.energy.gov/WebPAMSExternal/Interface/Awards/AwardSearchExternal.aspx'

    parser = argparse.ArgumentParser(description='Scrape DOE-funded awards')
    parser.add_argument('-s', '--start', dest='start_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-e', '--end', dest='end_date', help='range start date, format = YYYYMMDD', required=True)
    parser.add_argument('-u', '--userlist', dest='userlist', help='list of names and affiliations', required=True)
    parser.add_argument('-o', '--output', dest='output', help='output file', required=True)
    args = parser.parse_args()

    start = datetime.datetime.strptime(args.start_date, '%Y%m%d').strftime('%-m/%-d/%Y')
    start_validation = datetime.datetime.strptime(args.start_date, '%Y%m%d').strftime('%Y-%m-%d-00-00-00')
    end = datetime.datetime.strptime(args.end_date, '%Y%m%d').strftime('%-m/%-d/%Y')
    end_validation = datetime.datetime.strptime(args.end_date, '%Y%m%d').strftime('%Y-%m-%d-23-59-59')

    final_results = []  # A list of dictionaries, each containing the data of a single award entry

    make_requests(url, start, start_validation, end, end_validation, final_results)
    write_output_sheet(final_results, './data/' + args.userlist, './data/' + f'DOE_{args.output}')


def make_requests(url, start, start_validation, end, end_validation, final_results):
    """
    Retrieve specific award information by making several POST requests. The information returned
    is in the FINAL_RESULTS list.
    """
    with requests.Session() as session:
        # Start a session with a post request to the url
        try:
            res = session.post(url, timeout=20)
        except requests.exceptions.ReadTimeout:
            logging.error('timeout during search...try again later')
            sys.exit()
        except Exception as x:
            logging.error(f'request failed because {x}')
            sys.exit()

        # Use response to grab fields necessary for a valid search to go through
        soup = BeautifulSoup(res.content, 'html.parser')

        # Update payload with fields, incl. search params
        payload = {
            "ctl00_REIRadScriptManager1_TSM": soup.find(attrs={"name": "ctl00_REIRadScriptManager1_TSM"})['value'],
            "__EVENTTARGET": "ctl00$MainContent$grdAwardsList",
            "__EVENTARGUMENT": "FireCommand:ctl00$MainContent$grdAwardsList$ctl36;PageSize;100",
            "__VIEWSTATE": soup.find(attrs={"name": "__VIEWSTATE"})['value'],
            "__VIEWSTATEGENERATOR": soup.find(attrs={"name": "__VIEWSTATEGENERATOR"})['value'],
            # Institution name like:
            "ctl00$MainContent$pnlSearch$txtInstitutionName": "University of Texas",
            # Award start date:
            "ctl00$MainContent$pnlSearch$dpPPSDFrom$dateInput": f"{start}",
            "ctl00_MainContent_pnlSearch_dpPPSDFrom_dateInput_ClientState":
                f"{{'enabled':true,'emptyMessage':'','validationText':'{start_validation}', \
                    'valueAsString':'{start_validation}','minDateStr':'1980-00-01-00-01-00', \
                    'maxDateStr':'2099-00-31-00-12-00','lastSetTextBoxValue':'{start}'}}",
            "ctl00$MainContent$pnlSearch$dpPPSDTo$dateInput": f"{end}",
            "ctl00_MainContent_pnlSearch_dpPPSDTo_dateInput_ClientState":
                f"{{'enabled':true,'emptyMessage':'','validationText':'{end_validation}', \
                    'valueAsString':'{end_validation}','minDateStr':'1980-00-01-00-01-00', \
                    'maxDateStr':'2099-00-31-00-12-00','lastSetTextBoxValue':'{end}'}}",
        }

        # Make another request to update results per page with __EVENTARGUMENT param
        try:
            res = session.post(url, data=payload)
        except Exception as x:
            logging.error(f'request failed because {x}')
            sys.exit()

        # Grab updated viewstate that includes larger results per page included
        # Update payload
        soup = BeautifulSoup(res.content, 'html.parser')
        payload['__VIEWSTATE'] = soup.find(attrs={"name": "__VIEWSTATE"})['value']

        # Finally, make first search
        try:
            res = session.post(url, data=payload)
        except Exception as x:
            logging.error(f'request failed because {x}')
            sys.exit()
        parse_html(res.content, final_results)

        # get all event target values
        event_target = soup.find_all("div", {"class": "rgNumPart"})[0]
        event_target_list = [
            re.search('__doPostBack\(\'(.*)\',', t["href"]).group(1)
            for t in event_target.find_all('a')
        ]

        # Make updated post request to perform actual search
        for link in event_target_list[1:]:
            payload['__EVENTTARGET'] = link
            payload['__VIEWSTATE'] = soup.find(attrs={"name": "__VIEWSTATE"})['value']
            try:
                res = session.post(url, data=payload)
            except Exception as x:
                logging.error(f'request failed because {x}')
                sys.exit()
            soup = BeautifulSoup(res.content, "html.parser")
            parse_html(res.content, final_results)


def parse_html(response_content, final_results):
    """
    Given a POST response HTML page, grab all search results (award listings)
    on the page and append their results to the final_results list.
    """
    # Grab fields that contain the data
    soup = BeautifulSoup(response_content, 'html.parser')
    table = soup.find(class_="rgMasterTable")
    tbody = table.contents[5]
    trs = tbody.find_all("tr")
    tr_heads, tr_bodies = [], []
    while trs:
        tr_heads.append(trs.pop(0))
        tr_bodies.append(trs.pop(0))
    logging.info(f"{len(tr_heads)} award entries found on this page")
    results_list = []

    # Grab data from fields
    for head in tr_heads:
        tds = head.find_all("td")
        results_list.append({
            'Award Number': tds[1].text.strip(),
            'Title': tds[2].text.strip(),
            'Institution': tds[3].text.strip(),
            'PI First Name': tds[4].text.strip().split(', ')[1],
            'PI Last Name': tds[4].text.strip().split(', ')[0]
        })
    for index, body in enumerate(tr_bodies):
        lis = body.find_all("li")
        listing = results_list[index]
        listing['Org Code'] = lis[0].text.strip().split(':')[1]
        listing['Program Office'] = lis[1].text.strip().split(':')[1]
        listing['PM'] = lis[2].text.strip().split(':')[1]
        listing['Start Date'] = lis[6].text.strip().split(':')[1]
        listing['End Date'] = lis[7].text.strip().split(':')[1]
        listing['Most Recent Award Date'] = lis[8].text.strip().split(':')[1]
        listing['Award Type'] = lis[9].text.strip().split(':')[1]
        listing['Amount Awarded to Date'] = lis[10].text.strip().split(':')[1]
        listing['Amount Awarded this FY'] = lis[11].text.strip().split(':')[1]
        listing['Institution Type'] = lis[12].text.strip().split(':')[1]
        listing['UEI'] = lis[13].text.strip().split(':')[1]
        listing['Program Area'] = lis[14].text.strip().split(':')[1]
        listing['Register Number'] = lis[15].text.strip().split(':')[1]
        listing['DUNS'] = lis[16].text.strip().split(':')[1]

    # Append to final_results list
    final_results += results_list


def write_output_sheet(award_dict, userlist, output):
    """
    Given a dictionary of award information and a list of TACC usernames, write
    an output workbook with two worksheets: (1) Awards that match a TACC username
    and (2) awards that don't match a TACC username.
    """

    userlist_wb = load_workbook(filename=userlist, read_only=True)
    worksheet = userlist_wb['utrc_institution_accounts']
    row_count = worksheet.max_row
    rows = worksheet.rows

    name_dict = {}

    if row_count > 1:
        next(rows)  # skip header row
        for row in rows:
            institution = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            name = ' '.join([first_name, last_name]).lower().replace(' ', '')
            name_dict[name] = [institution, first_name, last_name]

    logging.info(f'number of items in name_dict = {len(name_dict.keys())}')
    logging.info(f'number of awards found = {len(award_dict)}')

    workbook = xlsxwriter.Workbook(output)
    bold = workbook.add_format({'bold': 1})
    found_worksheet = workbook.add_worksheet('utrc_doe_funding')
    found_worksheet.write_row(0, 0, ['utrc_institution', 'utrc_first_name', 'utrc_last_name']+AWARD_INFO, bold)
    not_found_worksheet = workbook.add_worksheet('not_utrc_doe_funding')
    not_found_worksheet.write_row(0, 0, AWARD_INFO, bold)

    f_row = 1
    nf_row = 1
    # For each award found through this search
    for item in award_dict:
        base_info = [item['Award Number'],
                     item['Title'],
                     item['Institution'],
                     item['PI First Name'],
                     item['PI Last Name'],
                     item['Org Code'],
                     item['Program Office'],
                     item['PM'],
                     item['Start Date'],
                     item['End Date'],
                     item['Most Recent Award Date'],
                     item['Award Type'],
                     item['Amount Awarded to Date'],
                     item['Amount Awarded this FY'],
                     item['Institution Type'],
                     item['UEI'],
                     item['Program Area'],
                     item['Register Number'],
                     item['DUNS']]
        name_str = f"{item['PI First Name']}{item['PI Last Name']}".lower().replace(' ', '')

        # Add it to found if exact match
        if name_str in name_dict.keys():
            logging.info(f"{name_str} matches {name_dict[name_str]}")
            found_worksheet.write_row(f_row, 0,
                                      [name_dict[name_str][0],
                                       name_dict[name_str][1],
                                       name_dict[name_str][2]
                                       ]+base_info)
            f_row += 1
        # Otherwise, do a fuzzy pattern matching check
        else:
            fuzzycheck = fuzzy_match(item, name_dict, found_worksheet,
                                     not_found_worksheet, name_str,
                                     base_info, workbook, f_row, nf_row)

            if fuzzycheck is False:
                nf_row += 1
            else:
                f_row += 1

    workbook.close()
    return


def fuzzy_match(item, name_dict, found_worksheet,
                not_found_worksheet, name_str,
                base_info, workbook, f_row, nf_row):
    """
    Given a search results award PI name, fuzzy match against each entry in
    the TACC userlist database. Accept matches in two categories--80-88% and 89+%.
    """
    f_format = workbook.add_format({'bg_color': '#90EE90'})
    nf_format = workbook.add_format({'bg_color': '#FCC981'})
    for key, values in name_dict.items():
        # Last name MUST match exactly
        if item['PI Last Name'] == values[2]:
            match_percent = fuzz.ratio(item['PI First Name'], name_dict[key][1])
            # If first name 89+% match, add it to found and highlight green
            if match_percent >= 89:
                logging.info(f"{name_str} fuzzy matches {values[1:2]} \
                               --match percent = {match_percent}")
                found_worksheet.write_row(f_row, 0, [values[0], values[1], values[2]]
                                          + base_info
                                          + [f"match percent = {match_percent}"], f_format)
                return True
            # If first name 80+% match, add it to found and highlight orange
            elif match_percent >= 80:
                logging.info(f"{name_str} fuzzy matches {values[1:2]} \
                               --match percent = {match_percent}")
                # Log and color the field in orange
                found_worksheet.write_row(f_row, 0, [values[0], values[1], values[2]]
                                          + base_info
                                          + [f"match percent = {match_percent}"], nf_format)
                return True
    # If no matches >=80%, add to not found workbook
    logging.info(f"{item['PI First Name']}{item['PI Last Name']} has no match")
    not_found_worksheet.write_row(nf_row, 0, base_info)
    return False


main()
